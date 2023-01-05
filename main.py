import os
import shutil
import threading
import time
import pygame
import Simulation as sim
import sys
import GlobalData as GD
import FileController as fc
import PDFReport as pdf
from reportlab.pdfgen import canvas
import openpyxl
from DropdownMenu import DropdownMenu
import numpy as np

def run_thread(thread_name:str , thread_target,args=()):
    thread = threading.Thread(name=thread_name, target=thread_target, args=args)
    thread.daemon = True
    thread.start()
    return thread





screen = pygame.display.set_mode(GD.screen_size)
pygame.display.set_caption("SIMULATION")
font = pygame.font.Font(None, GD.font_size)

def turn_signal_on(intersection :int, signal_img, index:int):
    screen.blit(signal_img, GD.intersections[intersection].signal_coordinates[index])




def display_signal_timer_and_vehicle_count_for_each_signal(intersection :int, signal_number:int , signal_texts : list ):
        signal_texts[signal_number] = font.render(str(GD.intersections[intersection].signals[signal_number].signal_text), True, GD.white, GD.black)
        screen.blit(signal_texts[signal_number], GD.intersections[intersection].signal_timer_coordinates[signal_number])
        
        displayText = GD.crossed[intersection][GD.direction_numbers[signal_number]]['crossed']
        #displayText = GD.vehicles_[GD.direction_numbers[signal_number]]['crossed']
        GD.intersections[intersection].vehicle_count_texts[signal_number] = font.render(str(displayText), True, GD.black, GD.white)
        screen.blit(GD.intersections[intersection].vehicle_count_texts[signal_number], GD.intersections[intersection].vehicle_count_coordinates[signal_number])
        

def display_the_vehicles():
    for vehicle in sim.simulation:
        #screen.blit(vehicle.current_image, [vehicle.x, vehicle.y])
        vehicle.move_(screen)


def display_time_elapsed():
    algorithm_activity = 'OFF'
    if(GD.algorithm_active):
        algorithm_activity = 'ON'
    time_elapsed_text = font.render(f"Algorithm Activity Status  : {algorithm_activity} ", True, GD.black, GD.white)
    screen.blit(time_elapsed_text, (750, 20))
    time_elapsed_text = font.render(("Time Elapsed: " + str(GD.time_elapsed)), True, GD.black, GD.white)
    screen.blit(time_elapsed_text, (750, 50))
    


def signals_conroller(intersection):
            traffic_sign_arrow_images = []

            for dir in GD.direction_numbers.values():
                traffic_sign_arrow_images.append(pygame.image.load(f'images/traffic_signs/{dir}.png'))
            
            for i,coordinate in enumerate(GD.intersections[intersection].traffic_sign_arrow_coordinates):
                screen.blit(traffic_sign_arrow_images[i], (coordinate[0], coordinate[1]))



            for i in range(0, GD.intersections[intersection].number_of_signals):
                if(i == GD.intersections[intersection].current_green):
                    if(GD.intersections[intersection].current_yellow == 1):
                        GD.intersections[intersection].signals[i].signal_text = GD.intersections[intersection].signals[i].yellow
                        turn_signal_on(intersection ,signal_img=GD.yellow_signal_img, index=i)

                        
                    else:
                        GD.intersections[intersection].signals[i].signal_text = GD.intersections[intersection].signals[i].green
                        if(GD.intersections[intersection].signals[i].green <= 6 and GD.intersections[intersection].signals[i].green > 0 and GD.intersections[intersection].signals[i].green % 2 == 0):
                            turn_signal_on(intersection ,signal_img = GD.non_signal, index=i)
                            
                        elif(GD.intersections[intersection].signals[i].green <= 6 and GD.intersections[intersection].signals[i].green > 0 and GD.intersections[intersection].signals[i].green % 2 == 1):
                            turn_signal_on(intersection ,signal_img = GD.green_signal, index=i)
                        
                        else :
                            turn_signal_on(intersection ,signal_img = GD.green_signal, index=i)
        
                else:
                    # if(GD.intersections[intersection].signals[i].red  < 5 and i == GD.intersections[intersection].next_green):
                    #     GD.intersections[intersection].signals[i].signal_text = GD.intersections[intersection].signals[i].red 
                    # elif(i != GD.intersections[intersection].next_green or i != GD.intersections[intersection].current_green):
                    if(i != GD.intersections[intersection].current_green):
                        GD.intersections[intersection].signals[i].signal_text = ""
                    turn_signal_on(intersection , signal_img = GD.red_signal_img, index=i)
                

            signal_texts = ["", "", "", ""]
            for i in range(0, GD.intersections[intersection].number_of_signals):
                display_signal_timer_and_vehicle_count_for_each_signal(intersection ,signal_number = i , signal_texts = signal_texts)


path = ''
    
def output():
    global path
    # Check if the directory exists
    if not os.path.exists('temp'):
        # Create the directory
        os.makedirs('temp')
    path , current_time = fc.create_directory()
    fc.copy_file(src='configuration/Algorithm.xlsx' , dst=path)   
    fc.copy_file(src='configuration/vehicles_db.xlsx' , dst=path)   
    
    fc.create_xlsx_file(path)
    for vehicle in sim.simulation:
        data = { 
            'vehicle_type':vehicle.vehicle_class, 
            'vehicle_speed_avg':vehicle.speed_avg
            }
        fc.append_dict_to_xlsx( filename= 'vehicles_avg_speeds.xlsx' , data=data  ,path=path)
    if(len(sim.simulation) <= 0):
        data = { 
            'vehicle_type':"", 
            'vehicle_speed_avg':0
            }
        fc.append_dict_to_xlsx( filename= 'vehicles_avg_speeds.xlsx' , data=data ,path=path )
    # Plot the average speeds for the specified vehicle types
    fc.plot_average_speeds_for_each_vehicle_type(path)
    fc.plot_vehicle_average_speed(path)
    pdf.create_report_for_current_configuration(path , current_time)

    if(len(os.listdir(path.rsplit('/', 1)[0])) %2 == 0):
        fc.create_simulations_data(path)
        fc.plot_simulations_data(path)
        pdf.create_report_for_overall_configurations(path,current_time)
    

def to_percent(fraction):
    percent = int(round(fraction * 100))
    return f"{percent}%" , percent

def create_text(displayText , position, font_size,font_color):
    font = pygame.font.SysFont('Arial', 22)

    # Create the text surface
    text_surface = font.render(displayText, True, font_color)

    # Get the rectangle for the text surface
    text_rect = text_surface.get_rect()

    # Set the position of the text rectangle
    text_rect.center = position 

    # Draw the text to the screen
    screen.blit(text_surface, text_rect) 

def show_loading_report():
    temp =  GD.sim_time -5  
    while(True ):
        screen.blit(GD.background_white, (0, 0))
        displayText , percent = to_percent(( GD.time_elapsed - GD.sim_time ) / (temp))
        displayText = 100 + 2*percent
        create_text(f'Loading Report {displayText}%' , ( 560,70) ,50,GD.black)
        screen.blit(GD.report, (337, 125))
        pygame.display.update()


###############################################################################################################################
###############################################################################################################################
###############################################################################################################################

# Set the dimensions of the dropdown menu
menu_width = 75
menu_height = 25




# Create the options for the dropdown menu
options = ["On", "Off"]

def read_algorithm_activity():
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Algorithm.xlsx")
    sheet = workbook["Sheet1"]
    if( sheet.cell(row=1, column=2).value.lower()  == 'true'):
        return  options[0]
    return  options[1]



def read_simulation_time():
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Simulation.xlsx")
    sheet = workbook["Sheet1"]
    return str(sheet.cell(row=1, column=2).value),str(sheet.cell(row=2, column=2).value)
    

# Set the default option
default_option = read_algorithm_activity()

# Open the Excel workbook
workbook = openpyxl.load_workbook('configuration/vehicles_db.xlsx')

# Get the first worksheet
worksheet = workbook['Sheet1']

# Define some constants for the table layout
TABLE_TOP = 100
TABLE_LEFT = 75
TABLE_WIDTH = 1000
TABLE_HEIGHT = 600
CELL_WIDTH = TABLE_WIDTH // 5
CELL_HEIGHT = TABLE_HEIGHT // 6

font = pygame.font.SysFont('Arial', 22)
menu = DropdownMenu(menu_width, menu_height, options, default_option, font)

# Set the position of the dropdown menu
menu_x = TABLE_WIDTH-50-menu_width
menu_y = 40
menu.set_position(menu_x, menu_y)

# Create a surface to draw the table on
table_surface = pygame.Surface((TABLE_WIDTH, TABLE_HEIGHT))
# Define the input rectangle
input_rect = pygame.Rect(TABLE_LEFT+270,40, 75, 25)
input_string,sim_id = read_simulation_time()
text_surface = font.render(input_string, True, (255, 255, 255))
# Set up the input string




def create_overall_config_downloading_button(color = GD.gray_fatih):
    """Create the Save button and its surface."""
    global overall_config_downloading_button, overall_config_downloading_button_surface

    # Create the Save button
    overall_config_downloading_button = pygame.Rect(190, 450, CELL_WIDTH , 50)
    overall_config_downloading_button_text = font.render('Download', True, (255, 255, 255))

    # Create the surface to display the Save button on
    overall_config_downloading_button_surface = pygame.Surface((CELL_WIDTH, 50))
    overall_config_downloading_button_surface.fill(color)
    overall_config_downloading_button_surface.blit(overall_config_downloading_button_text, (CELL_WIDTH // 2 - overall_config_downloading_button_text.get_width() // 2, 25 - overall_config_downloading_button_text.get_height() // 2))



def create_current_config_downloading_button(color = GD.gray_fatih):
    """Create the Save button and its surface."""
    global current_config_downloading_button, current_config_downloading_button_surface

    # Create the Save button
    current_config_downloading_button = pygame.Rect(690, 450, CELL_WIDTH , 50)
    current_config_downloading_button_text = font.render('Download', True, (255, 255, 255))

    # Create the surface to display the Save button on
    current_config_downloading_button_surface = pygame.Surface((CELL_WIDTH, 50))
    current_config_downloading_button_surface.fill(color)
    current_config_downloading_button_surface.blit(current_config_downloading_button_text, (CELL_WIDTH // 2 - current_config_downloading_button_text.get_width() // 2, 25 - current_config_downloading_button_text.get_height() // 2))


# def create_reset_button():
#     """Create the Save button and its surface."""
#     global reset_button, reset_button_surface

#     reset_button = pygame.Rect(400, 78, CELL_WIDTH//2 , 25)
#     reset_button_text = font.render('Reset', True, (255, 255, 255))

#     reset_button_surface = pygame.Surface((CELL_WIDTH//2, 25))
#     reset_button_surface.fill(gray_dark)
#     reset_button_surface.blit(reset_button_text, (CELL_WIDTH // 4 - reset_button_text.get_width() // 2, reset_button_text.get_height() // 2 -10))




def create_save_button():
    """Create the Save button and its surface."""
    global save_button, save_button_surface

    # Create the Save button
    save_button = pygame.Rect(TABLE_LEFT, TABLE_TOP + TABLE_HEIGHT + 10, CELL_WIDTH , 50)
    save_button_text = font.render('Save', True, (255, 255, 255))

    # Create the surface to display the Save button on
    save_button_surface = pygame.Surface((CELL_WIDTH, 50))
    save_button_surface.fill(GD.gray_dark)
    save_button_surface.blit(save_button_text, (CELL_WIDTH // 2 - save_button_text.get_width() // 2, 25 - save_button_text.get_height() // 2))



def create_start_button():
    """Create the start button and its surface."""
    global start_button, start_button_surface

    # Create the Save button
    start_button = pygame.Rect(TABLE_LEFT+TABLE_HEIGHT+150, TABLE_TOP + TABLE_HEIGHT + 10, CELL_WIDTH , 50)
    start_button_text = font.render('Srart', True, (255, 255, 255))

    # Create the surface to display the Start button on
    start_button_surface = pygame.Surface((CELL_WIDTH, 50))
    start_button_surface.fill(GD.gray_dark)
    start_button_surface.blit(start_button_text, (CELL_WIDTH // 2 - start_button_text.get_width() // 2, 25 - start_button_text.get_height() // 2))






def draw_table():
    """Draw the table and its cells."""
    global table_surface

    # Draw the table
    table_surface.fill(GD.white)
    for row in range(5):
        for col in range(4):
            # Calculate the cell rect
            cell_rect = pygame.Rect(TABLE_LEFT + col * CELL_WIDTH, TABLE_TOP + row * CELL_HEIGHT, CELL_WIDTH, CELL_HEIGHT)
           
            # Draw the cell rect
            pygame.draw.rect(table_surface, GD.gray, cell_rect, 1)
            # Render the cell text
            text = table[row][col]
            cell_text = font.render((str(text)).encode('utf-8'), True, GD.gray)
            # Calculate the position of the cell text
            cell_text_pos = (TABLE_LEFT + col * CELL_WIDTH + CELL_WIDTH // 2 - cell_text.get_width() // 2, TABLE_TOP + row * CELL_HEIGHT + CELL_HEIGHT // 2 - cell_text.get_height() // 2)
            # Draw the cell text
            table_surface.blit(cell_text, cell_text_pos)



def read_sim_numbers():
    path = f"configuration/Simulation.xlsx"
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Sheet1"]

    # Update the value of the cell
    return sheet.cell(row=2, column=2).value 
    

def set_sim_numbers(number):
    path = f"configuration/Simulation.xlsx"
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Sheet1"]

    # Update the value of the cell
    sheet.cell(row=2, column=2).value = f"{number}"
    workbook.save(path)
    

def save_data(value , filename):
    path = f"configuration/{filename}.xlsx"
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Sheet1"]

    # Define the new column name
    new_column_name = f"{value}"
    # Update the value of the cell
    sheet.cell(row=1, column=2).value = new_column_name
    if(filename == 'Simulation'):
        sheet.cell(row=2, column=2).value = '0'

    try:

         # Save the changes to the workbook
        workbook.save(path)
        print(f"successfully data saved on {filename}.xlsx")
        return True
    except Exception as e:
        print(f"Error saving {filename}.xlsx data : {e}")
        return False

# Create a lock object
lock = threading.Lock()

def handle_save_message(msg , color):
    # Acquire the lock
    lock.acquire()
    try:
        timer = 500
        
        while(timer > 0):
            # Create the font object with the specified size
            font = pygame.font.Font(None, 30)

            # Create the text surface
            text_surface = font.render(f"                                                                         {msg}                                                                         ", True, (255,255,255),color)

            # Get the rectangle for the text surface
            text_rect = text_surface.get_rect()

            # Set the position of the text rectangle
            text_rect.center = (550,10) 

            # Draw the text to the screen
            screen.blit(text_surface, text_rect)
            timer -=1
    finally:
        # Release the lock when finished
        lock.release()

def append_failed_field(failed_data , field_name):
    if(len(failed_data) == 0):
        failed_data += field_name
    else:
        failed_data += f', {field_name}'
    return failed_data

def save_table_data():
    workbook = openpyxl.load_workbook('configuration/vehicles_db.xlsx')
    sheet = workbook["Sheet1"]
    max_vehicles = 48
    vehicles_counter = 0
    for row in range(2, 6):
        vehicles_counter += int(table[row-1][3])


    path = ''
    generating_number_error = -1
    if(vehicles_counter < max_vehicles):
        """Save the values in the table to the Excel workbook."""
        for row in range(1, 6):
            for col in range(1, 5):
                # Set the cell value in the worksheet
                print(table[row-1][col-1])
                sheet.cell(row=row, column=col, value=table[row-1][col-1])
        path = 'configuration/vehicles_db.xlsx'
        generating_number_error = 0
        
    try:
        # Save the changes to the workbook
        print(path)
        workbook.save(path)
        print(f"successfully data saved on vehicles_db.xlsx")
        return 1
    except Exception as e:
        print(f"Error saving vehicles_db.xlsx data : {e}")
        return generating_number_error



def save_table():
    failed_data = ""
    active = False
    if(menu.selected_option == 'On'):
        active = True

    vehicles_db_saved        = save_table_data()
    algorithm_activity_saved = save_data(value = active            , filename = 'Algorithm')
    simulation_time_saved    = save_data(value = int(input_string) , filename = 'Simulation')
    
    
    if(vehicles_db_saved != 1) :
        failed_data = append_failed_field(failed_data , field_name='Vehicles Table')

    if(not algorithm_activity_saved):
        failed_data = append_failed_field(failed_data , field_name='Algorithm Status Activity')
       
    if(not simulation_time_saved):
        failed_data = append_failed_field(failed_data , field_name='Simulation Time')



    if(algorithm_activity_saved and simulation_time_saved and vehicles_db_saved == 1):
        run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=("Your changes have been saved!",(111,149,111)))
    else:
        if( vehicles_db_saved == -1):
            run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"Your overall generating numbers exeeds the max=48. Please make a change and try again.",(255,99,71)))
        else:
            run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"Failed to save changes for : {failed_data}!",(255,99,71)))
   





# Initialize the table with the values from the worksheet
table = []
for row in worksheet.iter_rows():
    table_row = []
    for cell in row:
        table_row.append(cell.value)
    table.append(table_row)

# Create the Save button and its surface
#create_reset_button()
create_save_button()
create_start_button()
create_overall_config_downloading_button()
create_current_config_downloading_button()

background_white = pygame.image.load('images/bg-white.png')
# Run the game loop
def init():
    global input_string,sim_id
    running = True
    input_enabled = False
    while running:
        screen.blit(background_white, (0, 0))
        # Handle events
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
                pygame.quit()

            
            elif event.type == pygame.MOUSEBUTTONDOWN:
                input_enabled = False
                if menu.is_mouse_selection(event.pos):
                    menu.toggle_menu()
                else:
                    menu.set_menu_selection(event.pos)
                    
                if input_rect.collidepoint(event.pos):
                    # Handle mouse button down events
                    if event.button == 1:
                        # Left mouse button clicked
                        # Enable input
                        input_enabled = True
                    # elif event.button == 3:
                    else:
                        # Right mouse button clicked
                        # Disable input
                        input_enabled = False
                # Check if the Save button was clicked
                elif save_button.collidepoint(event.pos):
                    save_table()
                # elif reset_button.collidepoint(event.pos):
                #     sim_id = 0
                elif start_button.collidepoint(event.pos):
                    screen.fill((255, 255, 255))
                    running = False
                else:
                    # Check which cell was clicked
                    row = (event.pos[1] - TABLE_TOP) // CELL_HEIGHT
                    row -= 1
                    col = (event.pos[0] - TABLE_LEFT) // CELL_WIDTH
                    print(f"row {row }  col {col}")
                    if row > 0 and row < 5 and col > 0 and col < 4:
                        current_row = row
                        current_col = col
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_BACKSPACE:
                    if(input_enabled):
                        input_string = input_string[:-1]
                    else:
                        # Delete the character before the cursor in the current cell
                        table[current_row][current_col] = str(table[current_row][current_col])[:-1]
                elif event.unicode.isprintable():
                    number = ["0","1","2","3","4","5","6","7","8","9"]
                    if((event.unicode) in number and input_enabled):
                        if( len(input_string) <5):
                            input_string += event.unicode
                    else:
                        num = str(table[current_row][current_col])
                        
                        if(current_col == 1):
                            if "." not in num and ((event.unicode) in number or event.unicode == ".") and len(num) > 0 or ("." in num and (event.unicode) in number) or "." not in num and (event.unicode) in number and len(num) == 0:
                                num += event.unicode
                                table[current_row][current_col] = float(num)
                        else:
                            if((event.unicode) in number):
                                num += event.unicode
                                table[current_row][current_col] = int(num)
                    

                

        
        # Draw the table and the  buttons
        draw_table()
        # Draw the table surface on the window
        screen.blit(table_surface, (TABLE_LEFT, TABLE_TOP))
        screen.blit(save_button_surface, save_button)
        screen.blit(start_button_surface, start_button)
        #screen.blit(reset_button_surface, reset_button)
        
        create_text('Algorithm Activity Status : ', (menu_x-130,50) ,32,GD.gray)
        # Draw the dropdown menu
        menu.draw(screen)


        create_text('Simulation Time  : ', (253,50) ,32,GD.gray)
        # Update the text surface with the input string
        text_surface = font.render(input_string, True, GD.gray)
        # Draw the input rectangle to the screen
        pygame.draw.rect(screen, GD.gray_dark, input_rect, 1)
        # Draw the text surface to the screen
        screen.blit(text_surface, input_rect.topleft)
       
        create_text(f'Simulations NO. :  {sim_id}', (263,90) ,32,GD.gray)

        pygame.display.update()




def display_file_downloading_statuse(downloaded_path):
    running = True
    while running:
        if(os.path.exists(f"{downloaded_path}")):
            running = False
            run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=("Report Downloaded Successfully!",(111,149,111)))


###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
class Main:
    init()
    
    GD.sim_time        = fc.read_xlsx_file_for_sim()+5
    algorithm_activity = fc.read_xlsx_file_for_algo()
    if( algorithm_activity.lower() == 'true'):
        GD.algorithm_active = True

    GD.vehicles_generating  = fc.read_xlsx_file(directory = 'configuration' , filename = 'vehicles_db.xlsx',column='generating_number')
    GD.vehicles_weight      = fc.read_xlsx_file(directory = 'configuration' , filename = 'vehicles_db.xlsx',column='weight')
    GD.speeds               = fc.read_xlsx_file(directory = 'configuration' , filename = 'vehicles_db.xlsx',column='speed')
   
    cars_number:int = 0
    for v in GD.vehicles_generating.values():
        cars_number += v
    GD.cars_number = cars_number
    run_thread("generateVehicles" ,sim.generate_vehicle)
 
    while(GD.cars_number > 0):
        screen.blit(GD.loading, (0, 0))
        
        # Set the font and font size
        font = pygame.font.Font(None, 36)
        displayText,percent =to_percent((cars_number - GD.cars_number ) / cars_number)
        
        if(percent < 60):
            screen.blit(GD.red_signal_img_88, (350, 300))
        elif(percent >= 60 and percent < 90 ):
            create_text('Get Ready!' , ( 560,100) ,50,GD.white)
            screen.blit(GD.yellow_signal_img_88, (350, 300))
        elif(percent >= 90 ):
            create_text('Go!' , ( 580,100) ,50,GD.white)
            screen.blit(GD.green_signal_88 , (350, 300))

        
        create_text(displayText , ( 550,390) ,32,GD.white)
        
        pygame.display.update()   

    #########################################################################################################################
    ####################################################    Threads   #######################################################
    #########################################################################################################################
    run_thread(thread_name="simulationTime" ,thread_target=sim.simulation_time)
    run_thread("initialization" ,sim.initialize)
    running = True
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
        if(GD.time_elapsed == GD.sim_time-5):
            # t1  = run_thread("show_loading_report" , show_loading_report)
            t2  = run_thread("output" , output)
            # t1.join()
            # t2.join()
            running = False
            
        
       
    #########################################################################################################################
    #######################################      Display Background In Simulation     #######################################
    #########################################################################################################################
        screen.blit(GD.background_white, (0, 0))
        screen.blit(GD.background, (150, 0))   
        #mouse coordination
        # mousex, mousey = pygame.mouse.get_pos()
        # print(f"{mousex} , {mousey}")

    #########################################################################################################################
    #######################################  display signal and set timer according   #######################################
    #######################################  to current status: green, yello, or red  #######################################
    #########################################################################################################################
        run_thread(thread_name="FGKJ signals_conroller" ,thread_target=signals_conroller,args=(GD.FGKJ,)).join()
        run_thread(thread_name="NOSR signals_conroller" ,thread_target=signals_conroller,args=(GD.NOSR,)).join()
    
        
        display_time_elapsed()

        display_the_vehicles()
        pygame.display.update()



    sim_nums = int(read_sim_numbers())
    sim_nums += 1
    set_sim_numbers(str(sim_nums))

    current_config_downloading_is_available = False
    overall_config_downloading_is_available = False

    running = True
    while running:
        
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            elif event.type == pygame.MOUSEBUTTONDOWN:
                if current_config_downloading_button.collidepoint(event.pos):
                    if(current_config_downloading_is_available):
                        downloaded_path , can_download = pdf.download_file(path=path , report_name='report')
                        if(can_download):
                            run_thread(thread_name='display_file_downloading_statuse', thread_target= display_file_downloading_statuse,args=(downloaded_path,))
                     
                        else:
                            run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"The PDF file already exists in the downloads directory.",(255,99,71)))
                    else:
                        run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"File not available yet.",(255,99,71)))
                           
                elif overall_config_downloading_button.collidepoint(event.pos):
                    if(overall_config_downloading_is_available):
                        downloaded_path , can_download = pdf.download_file(path, report_name='simulations_report')
                        if(can_download):
                            run_thread(thread_name='display_file_downloading_statuse', thread_target= display_file_downloading_statuse,args=(downloaded_path,))
                     
                        else:
                            run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"The PDF file already exists in the downloads directory.",(255,99,71)))
                    else:
                        run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=(f"File not available yet.",(255,99,71)))
                      
                    
        # t1  = run_thread("show_loading_report" , show_loading_report)
        # t2  = run_thread("output" , output)
        # t1.join()
        # t2.join()
        screen.blit(GD.background_white, (0, 0))
        # displayText , percent = to_percent(( GD.time_elapsed - GD.sim_time ) / (temp))
        # displayText = 100 + 2*percent
        font = pygame.font.SysFont('Arial', 40)

        # Create the text surface
        text_surface = font.render('Loading Reports', True, GD.gray)

        # Get the rectangle for the text surface
        text_rect = text_surface.get_rect()

        # Set the position of the text rectangle
        text_rect.center = ( 560,70) 

        # Draw the text to the screen
        screen.blit(text_surface, text_rect) 
        
        #screen.blit(GD.report, (337, 125))
        x=230
        y=280
        screen.blit(GD.overall_config_downloading, (x, y))
        screen.blit(GD.current_config_downloading, (x+500, y))
        if(os.path.exists(f"{path}/report.pdf")):
            current_config_downloading_is_available = True
            create_current_config_downloading_button(GD.gray_dark)

        if(os.path.exists(f"{path}/simulations_report.pdf")):
            overall_config_downloading_is_available = True
            create_overall_config_downloading_button(GD.gray_dark)

        screen.blit(overall_config_downloading_button_surface, overall_config_downloading_button)
        screen.blit(current_config_downloading_button_surface, current_config_downloading_button)
        pygame.display.update()
