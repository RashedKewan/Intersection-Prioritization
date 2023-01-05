import os
import shutil
import matplotlib.pyplot as plt
import pandas as pd
import GlobalData as GD
import openpyxl
import datetime
import shutil
import Simulation as sim
import numpy as np



def copy_file(src ,dst):
  # Copy the file from the source to the destination
  shutil.copy(src, dst)



def get_current_time():
  now = datetime.datetime.now()
  date_time = now.strftime("%d-%m-%Y/%H-%M-%S")
  return date_time




def read_xlsx_file_for_sim(directory = 'configuration' , filename = 'Simulation.xlsx'):
  file_path = f"{directory}/{filename}"

  # Open the workbook
  workbook = openpyxl.load_workbook(file_path)

  # Get the sheet you want to access
  sheet = workbook['Sheet1']

  # Iterate over the columns in the sheet
  for i,col in enumerate(sheet.columns):
      # Find the column you want to access
      if i == 1:
          # Print the values in the column
          for cell in col:
            return int(cell.value)





def read_xlsx_file_for_algo(directory = 'configuration' , filename = 'Algorithm.xlsx'):
  file_path = f"{directory}/{filename}"

  # Open the workbook
  workbook = openpyxl.load_workbook(file_path)

  # Get the sheet you want to access
  sheet = workbook['Sheet1']

  # Iterate over the columns in the sheet
  for i,col in enumerate(sheet.columns):
      # Find the column you want to access
      if i == 1:
          # Print the values in the column
          for cell in col:
            status = 'OFF'
            if str(cell.value).lower() == 'true':
              status = 'ON '
            print('\n\n\n')
            print('------------------------------------')
            print(f"| Algorithm Activity Status  : {status} |")
            print('------------------------------------')
            print('\n\n\n')
            return cell.value





def read_xlsx_file(directory = 'configuration' , filename = 'vehicles_generating.xlsx' , column = 'generating_number'):
  file_path = f"{directory}/{filename}"
  xlSheet = "Sheet1"
  
  # Open the Excel file
  xlsx = pd.ExcelFile(file_path)
  
  # Check if the sheet name is valid
  if xlSheet not in xlsx.sheet_names:
    raise ValueError(f"Sheet '{xlSheet}' not found in file '{filename}'")
  
  # Load the data from the xlsx file
  df = pd.read_excel(xlsx, sheet_name=xlSheet)
  data = {}
  # Iterate over the rows of the dataframe
  for _, row in df.iterrows():
      # Access data for each column by column name
      vehicle_type = row['vehicle_type']
      generating_number = row[column]
      data[vehicle_type]=generating_number
  return data


# def read_xlsx_file(directory = 'configuration' , filename = 'vehicles_generating.xlsx'):
#   file_path = f"{directory}/{filename}"
#   xlSheet = "Sheet1"
#   # Load the data from the xlsx file
#   df = pd.read_excel(file_path,sheet_name = xlSheet)

#   # Iterate over the rows of the dataframe
#   for _, row in df.iterrows():
#       # Access data for each column by column name
#       vehicle_type = row['vehicle_type']
#       generating_number = row['generating_number']
#       GD.vehicles_generating[vehicle_type]=generating_number
      




def create_directory():
  current_time = get_current_time()
  path = f'output/{current_time}'
  os.makedirs(path, exist_ok=True)
  return path , current_time



def create_xlsx_file(path='temp'):
  # Create a new empty workbook
  wb = openpyxl.Workbook()
    
  # Get the active sheet
  sheet = wb.active
  print(sheet)
  # Modify the data in the sheet
  sheet['A1'] = 'vehicle_type'
  sheet['B1'] = 'vehicle_speed_avg'
  
  # Save the workbook to an xlsx file
  wb.save(f'{path}/vehicles_avg_speeds.xlsx')








def append_dict_to_xlsx( filename , data ,path='temp'):
  if(len(data) > 1):
    df = pd.read_excel(f"{path}/{filename}")

    # Append the new row to the dataframe
    df = df.append(data, ignore_index=True)

    # Write the data back to the xlsx file
    df.to_excel(f"{path}/{filename}", index=False)









def remove_directory(directory="output"):
  # Check if the directory exists
  if os.path.exists(directory):
      # Delete the directory and all of its contents
      shutil.rmtree(directory)
      print("Directory deleted")
  else:
      # Print an error message
      print("Directory does not exist")








def plot_vehicle_average_speed(path='temp'):

  file_path = f"{path}/vehicles_avg_speeds.xlsx"
  xlSheet = "Sheet1"
  # Open the Excel file
  xlsx = pd.ExcelFile(file_path)
  
     # Check if the sheet name is valid
  if xlSheet not in xlsx.sheet_names:
    raise ValueError(f"Sheet '{xlSheet}' not found in file '{file_path}'")
  # Load the data from the xlsx file
  df = pd.read_excel(file_path,sheet_name = xlSheet)
  
  colors = ['skyblue' , 'g' , 'r' , 'pink', 'coral']
  # Create a mapping from vehicle type to integer index
  vehicle_type_to_index = {'car': 0, 'bus': 1, 'truck': 2, 'motorcycle': 3}

  # Use the map method to transform the vehicle_type column
  df['vehicle_type_index'] = df['vehicle_type'].map(vehicle_type_to_index)
  if(len(sim.simulation) > 0):
  # Plot the scatter plot using the transformed column
    df.plot(kind="scatter", x="vehicle_type", y="vehicle_speed_avg", color=[colors[int(x)] for x in df['vehicle_type_index']])

  # Customize the appearance of the plot
  plt.xlabel("Vehicle Type")
  plt.ylabel("Average Speed (km/h)")
  plt.title("Average Speeds for Different Vehicle Types")

  plt.savefig(f"{path}/avg_speed_for_each_vehicle.png", dpi=300)




def plot_average_speeds_for_each_vehicle_type(path='temp'):
  
    file_path = f"{path}/vehicles_avg_speeds.xlsx"
    xlSheet = "Sheet1"
    # Load the data from the xlsx file
    
     # Open the Excel file
    xlsx = pd.ExcelFile(file_path)
  
     # Check if the sheet name is valid
    if xlSheet not in xlsx.sheet_names:
      raise ValueError(f"Sheet '{xlSheet}' not found in file '{file_path}'")
    df = pd.read_excel(file_path,sheet_name = xlSheet)
    # Create a scatter plot of the data
    df.plot(kind="scatter",x="vehicle_type", y="vehicle_speed_avg")
    if(len(sim.simulation)>0):
    # Group the rows by vehicle type and calculate the mean average speed
      mean_speeds = df.groupby("vehicle_type").mean()
      print(mean_speeds)
    # # Create a bar plot of the mean average speeds
      mean_speeds.plot(kind="bar", cmap='viridis')
      mean_speeds.to_excel(f'{path}/mean_speeds.xlsx', sheet_name='Sheet1')

    
    # Get the x-coordinates of the bar plot
      x_coords = list(range(len(mean_speeds)))
    
      for i, (index, row) in enumerate(mean_speeds.iterrows()):
        plt.text(x_coords[i]-0.25, row["vehicle_speed_avg"]+0.5, f"{row['vehicle_speed_avg']:.1f}", fontsize=10, ha="center")

    # Customize the appearance of the plot
    plt.xlabel("Vehicle Type")
    plt.ylabel("Average Speed (km/h)")
    plt.title("Mean Average Speeds for Different Vehicle Types")
    
    plt.savefig(f"{path}/average_speeds_for_each_type.png", dpi=300)






def create_simulations_data(path='temp'):
    
  # Create a new workbook
  workbook = openpyxl.Workbook()

  # Get the active worksheet
  worksheet = workbook.active
  worksheet.title  = 'Sheet1'
  # Set the column names
  worksheet.append(['vehicle_type', 'vehicle_speed_avg', 'generating_number', 'algorithm_is_active'])

  # Iterate through the directories in the 'output' directory
  for directory_date in os.listdir('output'):
    for directory_time in os.listdir(f'output/{directory_date}'):
        # Read the 'mean_speeds.xlsx' file in the current directory
        mean_speeds_file = openpyxl.load_workbook(f'output/{directory_date}/{directory_time}/mean_speeds.xlsx')
        mean_speeds_worksheet = mean_speeds_file.active

        # Read the 'configuration/vehicles_db.xlsx' file
        vehicles_db_file = openpyxl.load_workbook(f'output/{directory_date}/{directory_time}/vehicles_db.xlsx')
        vehicles_db_worksheet = vehicles_db_file.active

        # Read the 'configuration/Algorithm.xlsx' file
        algorithm_file = openpyxl.load_workbook(f'output/{directory_date}/{directory_time}/Algorithm.xlsx')
        algorithm_worksheet = algorithm_file.active

        # Iterate through the rows in the 'mean_speeds.xlsx' worksheet
        for row in mean_speeds_worksheet.iter_rows(min_row=2):
            # Get the values from the 'mean_speeds.xlsx' file
            vehicle_type = row[0].value
            vehicle_speed_avg = row[1].value

            # Find the generating_number in the 'configuration/vehicles_db.xlsx' file
            generating_number = None
            for db_row in vehicles_db_worksheet.iter_rows(min_row=1):
                if db_row[0].value == vehicle_type:
                    generating_number = db_row[3].value
                    break

            # Get the value from the 'configuration/Algorithm.xlsx' file
            algorithm_is_active = algorithm_worksheet.cell(row=1, column=2).value

            # Add the values to the 'simulations_data.xlsx' file
            worksheet.append([vehicle_type, vehicle_speed_avg, generating_number, algorithm_is_active])
    
    # Save the 'simulations_data.xlsx' file
    workbook.save(f'{path}/simulations_data.xlsx')



def plot_simulations_data(path='temp'):
  
    file_path = f"{path}/simulations_data.xlsx"
    xlSheet = "Sheet1"
    # Load the data from the xlsx file
    
     # Open the Excel file
    xlsx = pd.ExcelFile(file_path)
  
     # Check if the sheet name is valid
    if xlSheet not in xlsx.sheet_names:
      raise ValueError(f"Sheet '{xlSheet}' not found in file '{file_path}'")

    df = pd.read_excel(file_path,sheet_name = xlSheet)

    # Group the rows by generating_number and algorithm_is_active, and calculate the mean of the vehicle_type for each group
    grouped_df = df.groupby(["generating_number", "algorithm_is_active"],as_index=False).agg('mean')
    # Drop rows with NaN values from the grouped dataframe
    grouped_df = grouped_df.dropna()
    
    df = grouped_df
    df.to_excel(f"{path}/simulations_avg_speeds.xlsx", index=False)
    # Initialize lists
    labels = []
    Means_Using_Algorithm = []
    Means_Without_Algorithm = []

    # Iterate through the rows of the DataFrame
    for i, row in df.iterrows():
        # Extract values from DataFrame
        generating_number = row['generating_number']
        algorithm_is_active = row['algorithm_is_active']
        vehicle_speed_avg = row['vehicle_speed_avg']
       
        
        # Add vehicle speed avg to appropriate list
        if algorithm_is_active:
            
          # Add generating number to labels list
          labels.append(generating_number)
          Means_Using_Algorithm.append(vehicle_speed_avg)
        else:
          Means_Without_Algorithm.append(vehicle_speed_avg)

    # Set up bar chart
    x = np.arange(len(labels))  # the label locations
   
    width = 0.1  # the width of the bars

    fig, ax = plt.subplots()
    rects1 = ax.bar(x - width/2, Means_Using_Algorithm, width, label='Using_Algorithm')
    rects2 = ax.bar(x + width/2, Means_Without_Algorithm, width, label='Without_Algorithm')

    # Add labels, title, and custom x-axis tick labels
    ax.set_ylabel('Vehicles Speed AVG')
    ax.set_title('Vehicles Speed AVG grouped by algorithm using for each generating vehicles number', fontsize = 9)
    #ax.set_xticks(x, labels)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)

    ax.legend()

    # Add labels to bars
    for rect in rects1:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., height,'%.2f' % height,ha='center', va='bottom',fontsize=5)

    for rect in rects2:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., height,'%.2f' % height,ha='center', va='bottom',fontsize=5)

    fig.tight_layout()
    plt.savefig(f"{path}/simulations_analysis.png", dpi=300)

