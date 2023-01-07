import subprocess
import openpyxl
import GlobalData as GD 

def algorithm_is_active(is_active : bool = False):
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Algorithm.xlsx")
    sheet = workbook["Sheet1"]

    # Define the new column name
    new_column_name = f"{is_active}"
    
    # Update the value of the cell
    sheet.cell(row=1, column=2).value = new_column_name

    # Save the changes to the workbook
    workbook.save("configuration/Algorithm.xlsx")





def apply(is_active : bool = False,times = 5):
    # Activate Algorithm
    algorithm_is_active(is_active)    
    
    # Prepare options for generation_number in configuration/vehicles_db.xlsx file
    generating_vehicles_tests = [
        [1,0,0,0],
        [0,1,0,0],
        [0,0,1,0],
        [0,0,0,1]
    ]

    for i in range(2,13):
         generating_vehicles_tests.append([i,i,i,i])

    for test_values in generating_vehicles_tests:
        # Open the workbook and sheet
        workbook = openpyxl.load_workbook("configuration/vehicles_db.xlsx")
        sheet = workbook["Sheet1"]
        # Loop through the rows in the sheet and update the value in the "generating_number" column
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True)):
            if i >= len(test_values):
                break
            cell = sheet.cell(row=i+2, column=4)
            cell.value = test_values[i]
        # Save the changes to the workbook
        workbook.save("configuration/vehicles_db.xlsx")

        # run the program 'times' times for each option
        for i in range(times):
            subprocess.run(["python", "main.py"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

apply(is_active=False,times = 1)
apply(is_active=True,times = 1)

