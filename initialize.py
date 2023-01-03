import threading
import time
import openpyxl
import pygame
import subprocess
pygame.init() 


# Colours
black = (0, 0, 0)
white = (255, 255, 255)

# Screensize
screen_width:int = 1100
screen_height:int = 800 #800
screen_size = (screen_width, screen_height)
# Create the window and set its size
window = pygame.display.set_mode(screen_size)
pygame.display.set_caption("SIMULATION")


# Set the dimensions of the dropdown menu
menu_width = 75
menu_height = 40




# Create the options for the dropdown menu
options = ["On", "Off"]

def read_algorithm_activity():
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Algorithm.xlsx")
    sheet = workbook["Sheet1"]
    if( sheet.cell(row=1, column=2).value.lower()  == 'true'):
        return  options[0]
    return  options[1]



def read_simulation_time():
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Simulation.xlsx")
    sheet = workbook["Sheet1"]
    return str(sheet.cell(row=1, column=2).value)
    

# Set the default option
default_option = read_algorithm_activity()

# Open the Excel workbook
workbook = openpyxl.load_workbook('configuration/vehicles_db.xlsx')

# Get the first worksheet
worksheet = workbook['Sheet1']

# Define some constants for the table layout
TABLE_TOP = 100
TABLE_LEFT = 75
TABLE_WIDTH = 1000
TABLE_HEIGHT = 600
CELL_WIDTH = TABLE_WIDTH // 5
CELL_HEIGHT = TABLE_HEIGHT // 6

class DropdownMenu:
    def __init__(self, width, height, options, default_option, font):
        self.width = width
        self.height = height
        self.options = options
        self.default_option = default_option
        self.font = font
        self.selected_option = default_option
        self.menu_open = False

    def set_position(self, x, y):
        self.x = x
        self.y = y

    def draw(self, screen):
        # Draw the dropdown box
         # Draw the dropdown box
        pygame.draw.rect(screen, (200, 200, 200), (self.x, self.y, self.width, self.height), 1)

        # Render the selected option
        text = self.font.render(self.selected_option, 1, (0, 0, 0))
        screen.blit(text, (self.x + 10, self.y + 10))

        # If the menu is open, draw the options
        if self.menu_open:
            y = self.y + self.height
            for option in self.options:
                # Render the option
                text = self.font.render(option, 1, (0, 0, 0))
                screen.blit(text, (self.x + 10, y + 10))
                # Increment the y position for the next option
                y += self.height

    def toggle_menu(self):
        self.menu_open = not self.menu_open

    def is_mouse_selection(self, pos):
        if pos[0] > self.x and pos[0] < self.x + self.width and pos[1] > self.y and pos[1] < self.y + self.height:
            return True
        return False

    def set_menu_selection(self, pos):
        if self.menu_open:
            y = self.y + self.height
            for option in self.options:
                if pos[1] > y and pos[1] < y + self.height:
                    self.selected_option = option
                    self.menu_open = False
                y += self.height






# Set the font and the font size
font = pygame.font.Font(None, 32)
# Create the dropdown menu
menu = DropdownMenu(menu_width, menu_height, options, default_option, font)

# Set the position of the dropdown menu
menu_x = TABLE_LEFT+TABLE_WIDTH-40-menu_width
menu_y = 25
menu.set_position(menu_x, menu_y)

font_size:int = 24
font = pygame.font.SysFont('Arial', 18)

# Create a font for the table cells
font = pygame.font.Font(None, font_size)

# Create a surface to draw the table on
table_surface = pygame.Surface((TABLE_WIDTH, TABLE_HEIGHT))



# Define the input rectangle
input_rect = pygame.Rect(TABLE_LEFT+290,40, 75, 25)

# Set up the font and text surface
font = pygame.font.Font(None, 32)
text_surface = font.render(read_simulation_time(), True, (255, 255, 255))
# Set up the input string
input_string = read_simulation_time()





def create_text(displayText , position, font_size,font_color):
    # Create the font object with the specified size
    font = pygame.font.Font(None, font_size)

    # Create the text surface
    text_surface = font.render(displayText, True, font_color)

    # Get the rectangle for the text surface
    text_rect = text_surface.get_rect()

    # Set the position of the text rectangle
    text_rect.center = position 

    # Draw the text to the screen
    window.blit(text_surface, text_rect)







def create_save_button():
    """Create the Save button and its surface."""
    global save_button, save_button_surface

    # Create the Save button
    save_button = pygame.Rect(TABLE_LEFT, TABLE_TOP + TABLE_HEIGHT + 10, CELL_WIDTH , 50)
    save_button_text = font.render('Save', True, (255, 255, 255))

    # Create the surface to display the Save button on
    save_button_surface = pygame.Surface((CELL_WIDTH, 50))
    save_button_surface.fill((0, 0, 255))
    save_button_surface.blit(save_button_text, (CELL_WIDTH // 2 - save_button_text.get_width() // 2, 25 - save_button_text.get_height() // 2))






def draw_table():
    """Draw the table and its cells."""
    global table_surface

    # Draw the table
    table_surface.fill(white)
    for row in range(5):
        for col in range(4):
            # Calculate the cell rect
            cell_rect = pygame.Rect(TABLE_LEFT + col * CELL_WIDTH, TABLE_TOP + row * CELL_HEIGHT, CELL_WIDTH, CELL_HEIGHT)
            # Draw the cell rect
            pygame.draw.rect(table_surface, black, cell_rect, 1)
            # Render the cell text
            text = table[row][col]
            cell_text = font.render((str(text)).encode('utf-8'), True, black)
            # Calculate the position of the cell text
            cell_text_pos = (TABLE_LEFT + col * CELL_WIDTH + CELL_WIDTH // 2 - cell_text.get_width() // 2, TABLE_TOP + row * CELL_HEIGHT + CELL_HEIGHT // 2 - cell_text.get_height() // 2)
            # Draw the cell text
            table_surface.blit(cell_text, cell_text_pos)






def set_algorithm_activity(is_active : bool = False):
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Algorithm.xlsx")
    sheet = workbook["Sheet1"]

    # Define the new column name
    new_column_name = f"{is_active}"
    # Update the value of the cell
    sheet.cell(row=1, column=2).value = new_column_name

   
    
    try:
         # Save the changes to the workbook
        workbook.save("configuration/Algorithm.xlsx")
        print("algorithm_activity saved successfully")
        return True
    except Exception as e:
        print("Error saving algorithm_activity:", e)
        return False


def set_simulation_time(sim_time:int = 15):
    # Open the workbook and sheet
    workbook = openpyxl.load_workbook("configuration/Simulation.xlsx")
    sheet = workbook["Sheet1"]

    # Define the new column name
    new_column_name = f"{sim_time}"
    # Update the value of the cell
    sheet.cell(row=1, column=2).value = new_column_name

   
    try:
         # Save the changes to the workbook
        workbook.save("configuration/Simulation.xlsx")
        print("simulation_time saved successfully")
        return True
    except Exception as e:
        print("Error saving simulation_time:", e)
        return False




def run_thread(thread_name:str , thread_target,args=()):
    thread = threading.Thread(name=thread_name, target=thread_target, args=args)
    thread.daemon = True
    thread.start()
    return thread

def handle_save_message(msg , color):
    timer = 500
    
    while(timer > 0):
        # Create the font object with the specified size
        font = pygame.font.Font(None, 30)

        # Create the text surface
        text_surface = font.render(f"                                                  {msg}                                                  ", True, (255,255,255),color)

        # Get the rectangle for the text surface
        text_rect = text_surface.get_rect()

        # Set the position of the text rectangle
        text_rect.center = (550,15) 

        # Draw the text to the screen
        window.blit(text_surface, text_rect)
        timer -=1
        


def save_table():
    vehicles_db_saved = False
    """Save the values in the table to the Excel workbook."""
    for row in range(1, 6):
        for col in range(1, 5):
            # Set the cell value in the worksheet
            worksheet.cell(row=row, column=col, value=table[row-1][col-1])
    
    try:
        # Save the changes to the workbook
        workbook.save('configuration/vehicles_db.xlsx')
        print("Workbook saved successfully")
        vehicles_db_saved = True
    except Exception as e:
        print("Error saving workbook:", e)
       
    active = False
    if(menu.selected_option == 'On'):
        active = True
    algorithm_activity_saved = set_algorithm_activity(is_active=active)
    simulation_time_saved = set_simulation_time(sim_time= int(input_string))
    if(algorithm_activity_saved and simulation_time_saved and vehicles_db_saved):
        run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=("Your changes have been saved!",(154,205,50)))
    else:
        run_thread(thread_name='handle_save_message', thread_target= handle_save_message,args=("Your changes failed to save!",(255,99,71)))
   





# Initialize the table with the values from the worksheet
table = []
for row in worksheet.iter_rows():
    table_row = []
    for cell in row:
        table_row.append(cell.value)
    table.append(table_row)

# Create the Save button and its surface
create_save_button()
background_white = pygame.image.load('images/bg-white.png')
# Run the game loop
running = True
input_enabled = False
while running:
    window.blit(background_white, (0, 0))
    # Handle events
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
            
        elif event.type == pygame.MOUSEBUTTONDOWN:
            if input_rect.collidepoint(event.pos):
                # Handle mouse button down events
                if event.button == 1:
                    # Left mouse button clicked
                    # Enable input
                    input_enabled = True
                elif event.button == 3:
                    # Right mouse button clicked
                    # Disable input
                    input_enabled = False
            if menu.is_mouse_selection(event.pos):
                menu.toggle_menu()
            else:
                menu.set_menu_selection(event.pos)
            print(menu.selected_option)
            # Check if the Save button was clicked
            if save_button.collidepoint(event.pos):
                save_table()
            else:
                # Check which cell was clicked
                row = (event.pos[1] - TABLE_TOP) // CELL_HEIGHT
                row -= 1
                col = (event.pos[0] - TABLE_LEFT) // CELL_WIDTH
                #print(f"row {row }  col {col}")
                if row > 0 and row < 5 and col > 0 and col < 4:
                    current_row = row
                    current_col = col
        elif event.type == pygame.KEYDOWN:
            if event.key == pygame.K_BACKSPACE:
                if(input_enabled):
                    input_string = input_string[:-1]
                else:
                    # Delete the character before the cursor in the current cell
                    table[current_row][current_col] = str(table[current_row][current_col])[:-1]
            elif event.unicode.isprintable():
                number = ["0","1","2","3","4","5","6","7","8","9"]
                if((event.unicode) in number and input_enabled):
                    if( len(input_string) <5):
                        input_string += event.unicode
                else:
                    num = str(table[current_row][current_col])
                    
                    if(current_col == 1):
                        if "." not in num and ((event.unicode) in number or event.unicode == ".") and len(num) > 0 or ("." in num and (event.unicode) in number) or "." not in num and (event.unicode) in number and len(num) == 0:
                            num += event.unicode
                            table[current_row][current_col] = float(num)
                    else:
                        if((event.unicode) in number):
                            num += event.unicode
                            table[current_row][current_col] = int(num)
                

               

    
    # Draw the table and the Save button
    draw_table()
    window.blit(save_button_surface, save_button)
    # Draw the table surface on the window
    window.blit(table_surface, (TABLE_LEFT, TABLE_TOP))
    create_text('Algorithm Activity Status : ', (menu_x-150,50) ,32,(0,0,0))
    # Draw the dropdown menu
    menu.draw(window)


    create_text('Simulation Time : ', (250,50) ,32,(0,0,0))
    # Update the text surface with the input string
    text_surface = font.render(input_string, True, (0,0,0))
    # Draw the input rectangle to the screen
    pygame.draw.rect(window, (200, 200, 200), input_rect, 1)
    # Draw the text surface to the screen
    window.blit(text_surface, input_rect.topleft)

    pygame.display.update()

# Quit pygame
pygame.quit()
subprocess.run(["python", "main.py"])