import os
import PyPDF2
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen.canvas import Canvas
import GlobalData as GD





# Function to convert a table from an Excel file to an image
def excel_to_image(filename,path):
    # Open the Excel file
    wb = load_workbook(f'{path}/{filename}')
    ws = wb.active

    # Select the table
    table = []
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        table.append([cell.value for cell in row])

    # Convert the table to an image
    font        = ImageFont.truetype("arial.ttf", 12)
    img_width   = 500
    img_height  = len(table) * 20
    image       = Image.new("RGB", (img_width, img_height), (255, 255, 255))
    draw        = ImageDraw.Draw(image)
    y_offset    = 0

    for row in table:
        x_offset = 0
        for cell in row:
            draw.text((x_offset, y_offset), str(cell), font=font, fill=(0, 0, 0))
            x_offset += 150
        y_offset += 20

    # Save the image
    img = filename.replace(".xlsx", ".png")
    img = img.rsplit("/", 1)[-1]
    image.save(f"{path}/{img}")
    return img_height










def create_report_for_overall_configurations(path,current_time):
    # Convert the tables from the Excel files to images
    simulations_avg_speeds_image_height = excel_to_image(filename='simulations_avg_speeds.xlsx',path=path)

    #Create a new PDF document with a size of letter paper
    pdf_file = f"{path}/simulations_report.pdf"
    pdf_canvas = Canvas(pdf_file, pagesize=letter)
    
    y = 650
    # Title
    pdf_canvas.drawImage("images/braude_logo.png", 125, y, 400, 110)
    # Set the font
    pdf_canvas.setFont("Helvetica", 24)
    y = y -100
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(120, y, "Prioritize The Passage Of Vehicles At")
    y = y - 25
    pdf_canvas.drawString(150, y, "Intersections According To Fuel")
    y = y - 30
    pdf_canvas.drawString(250, y, "Consumption")
    y = y - 30
    pdf_canvas.setFillColorRGB(1, 0, 0)
    pdf_canvas.drawString(265, y, "-  Report  -")
    y = y - 100
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 16)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, "  Supervisor:  ")
    y = y - 20
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, " Mr. Ronen Zilber  ")
    y = y - 100
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 16)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, "     Authors:  ")
    y = y - 20
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(165, y, "Rashed Kewan 318369170 Rashed.Kewan@e.braude.ac.il")
    y = y - 20
    pdf_canvas.drawString(165, y, "Ibrahim Qassem 207925926 Ibrahim.Kassem@e.braude.ac.il")


    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    # Date
    time = current_time.rsplit("/", 1)[1]
    time = time.replace('-',':')
    date = current_time.rsplit("/", 1)[0]
    date = date.replace("-","/")
    y=100
    pdf_canvas.drawString(50, y, f"Date : {date}")
    pdf_canvas.drawString(50, y-15, f"Time : {time}")


   ###########################################################################################
    pdf_canvas.showPage()
    y = 750
    pdf_canvas.drawString(50, y, "__________________________    Vehicles Average Speeds    _______________________")
    y = y-25 - simulations_avg_speeds_image_height
    pdf_canvas.drawImage(f"{path}/simulations_avg_speeds.png", 50, y, 500, simulations_avg_speeds_image_height)


    ###########################################################################################
    pdf_canvas.showPage()
    # Add the plot image to the PDF document
    y = 750
    pdf_canvas.drawString(100, y, "___________________________    Graphs    ___________________________")
    pdf_canvas.drawImage(f"{path}/simulations_analysis.png", x=100, y=410, width=400, height=300)
    
    pdf_canvas.save()
    
        
    if os.path.exists(f"{path}/simulations_report.pdf"):
        print ('-----------------------------------------------')
        print(f"|   Simulations Report Successfully Created    |")
        print ('-----------------------------------------------')
        
    else:
        print ('-----------------------------------------------')
        print(f"|     Failed Simulations Report Creation.      |")
        print ('-----------------------------------------------')









def create_report_for_current_configuration(path , current_time):
    # Convert the tables from the Excel files to images
    vehicles_data_image_height = excel_to_image(filename='vehicles_db.xlsx',path=path)
    vehicles_avg_speeds_image_height = excel_to_image(filename='vehicles_avg_speeds.xlsx',path=path)

    #Create a new PDF document with a size of letter paper
    pdf_file = f"{path}/report.pdf"
    pdf_canvas = Canvas(pdf_file, pagesize=letter)
    
    y = 650
    # Title
    pdf_canvas.drawImage("images/braude_logo.png", 125, y, 400, 110)
    # Set the font
    pdf_canvas.setFont("Helvetica", 24)
    y = y -100
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(120, y, "Prioritize The Passage Of Vehicles At")
    y = y - 25
    pdf_canvas.drawString(150, y, "Intersections According To Fuel")
    y = y - 30
    pdf_canvas.drawString(250, y, "Consumption")
    y = y - 30
    pdf_canvas.setFillColorRGB(1, 0, 0)
    pdf_canvas.drawString(265, y, "-  Report  -")
    y = y - 100
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 16)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, "  Supervisor:  ")
    y = y - 20
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, " Mr. Ronen Zilber  ")
    y = y - 100
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 16)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(270, y, "     Authors:  ")
    y = y - 20
    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    pdf_canvas.drawString(165, y, "Rashed Kewan 318369170 Rashed.Kewan@e.braude.ac.il")
    y = y - 20
    pdf_canvas.drawString(165, y, "Ibrahim Qassem 207925926 Ibrahim.Kassem@e.braude.ac.il")


    
    # Set the font
    pdf_canvas.setFont("Helvetica", 12)
    # Set the fill color to black
    pdf_canvas.setFillColorRGB(0, 0, 0)
    # Date
    time = current_time.rsplit("/", 1)[1]
    time = time.replace('-',':')
    date = current_time.rsplit("/", 1)[0]
    date = date.replace("-","/")
    y=100
    pdf_canvas.drawString(50, y, f"Date : {date}")
    pdf_canvas.drawString(50, y-15, f"Time : {time}")


    ###########################################################################################
    pdf_canvas.showPage()
    y= 750
    # Algorithm
    algorithm_activity = 'OFF'
    if( GD.algorithm_active ):
        algorithm_activity = 'ON'
    pdf_canvas.drawString(50, y   , f"Algorithm Activity Status : {algorithm_activity}")
    pdf_canvas.drawString(50, y-15, f"Time Elapsed                  : {GD.sim_time-5}")
    

    y = y - 30
    
    # excel files
    pdf_canvas.drawString(50, y, "_____________________________    Vehicles Data    _____________________________")
    y = y - 150
    pdf_canvas.drawImage(f"{path}/vehicles_db.png", 50, y, 480, vehicles_data_image_height)

    
   ###########################################################################################
    pdf_canvas.showPage()
    y = 750
    pdf_canvas.drawString(50, y, "__________________________    Vehicles Average Speeds    _______________________")
    y = y-25 - vehicles_avg_speeds_image_height
    pdf_canvas.drawImage(f"{path}/vehicles_avg_speeds.png", 50, y, 500, vehicles_avg_speeds_image_height)



    ###########################################################################################
    pdf_canvas.showPage()
    # Add the plot image to the PDF document
    y = 750
    pdf_canvas.drawString(100, y, "___________________________    Graphs    ___________________________")
    pdf_canvas.drawImage(f"{path}/average_speeds_for_each_type.png", x=100, y=410, width=400, height=300)
    pdf_canvas.drawImage(f"{path}/avg_speed_for_each_vehicle.png", x=100, y=100, width=400, height=300)
    pdf_canvas.showPage()
    pdf_canvas.save()
    
        
    if os.path.exists(f"{path}/report.pdf"):
        print ('------------------------------------')
        print(f"|   Report Successfully Created    |")
        print ('------------------------------------')
        
    else:
        print ('------------------------------------')
        print(f"|     Failed Report Creation.      |")
        print ('------------------------------------')

def read_pdf_file(path):

    # Construct the path to the PDF file
    pdf_path = path

    # Open the PDF file in read-only mode
    with open(pdf_path, "rb") as f:
        # Create a PDF object
        pdf = PyPDF2.PdfReader(f)
            
        # Create a new PDF file to write to
        output_pdf = PyPDF2.PdfWriter()

        # Get the number of pages in the PDF
        num_pages = len(pdf.pages)
       
        # Iterate through the pages of the PDF
        for i in range(num_pages):
            # Get the i-th page
            page = pdf.pages[i]

            # Extract the text from the page
            output_pdf.add_page(page)

        
        return output_pdf

def download_file(path,report_name = 'report'):
    # Get the path to the downloads directory
    downloads_dir = os.path.expanduser("~/Downloads")

    time = path.rsplit("/")[-1]
    date = path.rsplit("/")[-2]
    # Construct the path to the PDF file in the downloads directory
    pdf_path = os.path.join(downloads_dir, f"{report_name}_{date}_{time}.pdf")

    # Check if the file already exists
    if os.path.exists(pdf_path):
        # The file already exists, so you can decide what action to take here
        print("The PDF file already exists in the downloads directory.")
        return pdf_path , False
    else:
        
        pdf_contents = read_pdf_file(f"{path}/{report_name}.pdf")
        # Create a new file object in write mode
        with open(pdf_path, "wb") as f:
            # Write the contents of the PDF to the file
            pdf_contents.write(f)

        # Close the file
        f.close()
        return pdf_path , True
